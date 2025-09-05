#!/usr/bin/env python3
"""
Script to examine the SAMPLE Excel file and understand its structure.
"""

import pandas as pd
import openpyxl

def examine_sample_excel():
    """Examine the SAMPLE Excel file structure."""
    print("Examining SAMPLE.xlsx structure:")
    print("="*50)
    
    # Read with pandas
    df = pd.read_excel('output/SAMPLE.xlsx')
    print(f"DataFrame shape: {df.shape}")
    print(f"Columns: {df.columns.tolist()}")
    
    # Read with openpyxl to get exact cell values
    wb = openpyxl.load_workbook('output/SAMPLE.xlsx')
    ws = wb.active
    
    print(f"\nWorksheet title: {ws.title}")
    print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
    
    print("\nFirst 15 rows with all columns:")
    print("-" * 100)
    
    for row_num in range(1, min(16, ws.max_row + 1)):
        row_data = []
        for col_num in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value is None:
                cell_value = ""
            row_data.append(str(cell_value)[:15])  # Limit length for display
        
        print(f"Row {row_num:2d}: {' | '.join(row_data)}")
    
    print("\nLooking for data patterns...")
    # Look for patterns in the data
    doc_ref_pattern = "P11569-11-99-40-2619"
    rows_with_doc_ref = []
    
    for row_num in range(1, min(ws.max_row + 1, 50)):
        for col_num in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_num, column=col_num).value
            if cell_value and doc_ref_pattern in str(cell_value):
                rows_with_doc_ref.append((row_num, col_num, str(cell_value)))
    
    print(f"\nFound {len(rows_with_doc_ref)} cells with document reference:")
    for row, col, value in rows_with_doc_ref[:10]:
        print(f"  Row {row}, Col {col}: {value}")

if __name__ == "__main__":
    examine_sample_excel()
